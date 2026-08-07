# job_audit_xlsx.py
from datetime import datetime, timezone #for generating UT timestamp if none given
from pathlib import Path  #for file path creation
import os #for os features like  build directory path

from openpyxl import Workbook, load_workbook #for creating/loading excel audit workbook
from openpyxl.utils import get_column_letter #for column sizes etc

HEADER = ["timestamp_utc", "fusion_alias", "job_name", "request_id", "status", "notes"]
SHEET_NAME = "Runs"

def _autosize(ws):
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

def _ensure_workbook(path: Path):
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADER)
        _autosize(ws)
        wb.save(path)
        return

    wb = load_workbook(path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    if ws.max_row == 0:
        ws.append(HEADER)
    else:
        row1 = [ws.cell(row=1, column=i+1).value for i in range(len(HEADER))]
        if row1 != HEADER:
            ws.insert_rows(1)
            for i, h in enumerate(HEADER, start=1):
                ws.cell(row=1, column=i, value=h)
    wb.save(path)

def _get_ws(path: Path):
    wb = load_workbook(path)
    return wb, (wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active)

def _existing_request_ids(ws):
    header = [ws.cell(row=1, column=i+1).value for i in range(len(HEADER))]
    req_col = header.index("request_id") + 1 if "request_id" in header else 4
    ids = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=req_col).value
        if v is not None:
            ids.add(str(v))
    return ids

def _existing_fallback_keys(ws):
    """
    Fallback key for rows without request_id:
    fusion_alias|job_name|minute_bucket (timestamp_utc truncated to minute)
    """
    header = [ws.cell(row=1, column=i+1).value for i in range(len(HEADER))]
    idx = {h: (header.index(h) + 1) for h in header if h in HEADER}
    keys = set()
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(row=r, column=idx["request_id"]).value if "request_id" in idx else None
        if rid:
            continue
        ts = ws.cell(row=r, column=idx["timestamp_utc"]).value if "timestamp_utc" in idx else ""
        minute_bucket = str(ts)[:16]  # YYYY-MM-DDTHH:MM
        fa = ws.cell(row=r, column=idx["fusion_alias"]).value if "fusion_alias" in idx else ""
        jn = ws.cell(row=r, column=idx["job_name"]).value if "job_name" in idx else ""
        keys.add(f"{fa}|{jn}|{minute_bucket}")
    return keys

def _default_base_dir():
    base = os.getenv("DATA_DROP_DIR", "").strip()
    return Path(base, "logs") if base else Path("logs")

def _default_xlsx_path():
    audit_dir = _default_base_dir()
    audit_dir.mkdir(parents=True, exist_ok=True)
    if os.getenv("JOB_AUDIT_DAILY", "").strip() in {"1", "true", "TRUE", "yes", "YES"}:
        day_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        return str(audit_dir / f"job_runs_{day_utc}.xlsx")
    return str(audit_dir / "job_runs.xlsx")

def log_job_run_xlsx(
    job_name: str,
    request_id: str,
    status: str,
    fusion_alias: str = "",
    notes: str = "",
    xlsx_path: str | None = None,
    timestamp_utc: str | None = None,  # NEW (optional ISO, e.g., '2025-10-09T14:03:22Z')
):
    """
    Append one row to the audit workbook; skip if request_id already exists.
    - If timestamp_utc is provided, use it; otherwise use current UTC time.
    - If request_id is empty, use a fallback de-dupe key (alias|job|minute).
    """
    xlsx_path = xlsx_path or _default_xlsx_path()
    p = Path(xlsx_path)
    _ensure_workbook(p)

    ts = (timestamp_utc or datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")).strip()

    try:
        wb, ws = _get_ws(p)

        req_id_str = ("" if request_id is None else str(request_id)).strip()
        if req_id_str:
            if req_id_str in _existing_request_ids(ws):
                wb.save(p)
                return
        else:
            minute_bucket = ts[:16]
            fallback_key = f"{fusion_alias}|{job_name}|{minute_bucket}"
            if fallback_key in _existing_fallback_keys(ws):
                wb.save(p)
                return

        ws.append([ts, fusion_alias, job_name, req_id_str, str(status), notes])
        _autosize(ws)
        wb.save(p)
    except PermissionError:
        # Workbook open/locked → ignore so your run doesn't fail
        pass
