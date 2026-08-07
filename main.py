# main.py
from __future__ import annotations  # must be the very first import, for simplifying type hinting throughout file

import os, json, time, re, shutil, argparse  #os for os features like reading env variables or file locations, json for parse/create schedule jsons, time for time related
# re for regular expressions, shutil for copying Excel files into main, argparse for simplifying reading command lines

from typing import Dict, List, Union  #typing for type hints and reading improvability
from pathlib import Path     #pathlib/path for handling file/directory paths

from crypto_env import load_env_from_encrypted, get_env, _get_env_prefix # dotenv for loading env variables from .env file into python but from the encypted version
from openpyxl import load_workbook  #openpyx1 for reading and writing Excel files

from logger import get_logger  # custom console+file logger
from fusion_api import ESSClient, ESSDuplicatePendingError   #custom that contains ESSClient class for communiction with Oracle Rest API
from job_audit_xlsx import log_job_run_xlsx  # for Excel logging of job runs
import sys
# setup/configuration
load_env_from_encrypted()   # load .env before reading env vars from encrypted version
    
ENV_PREFIX = _get_env_prefix()
logger = get_logger(__name__, ENV_PREFIX) # one logger for this module
# Stable paths
PROJECT_ROOT = Path(__file__).resolve().parent
DATA_XLSX_PATH = Path(os.getenv("DATA_XLSX_PATH") or (PROJECT_ROOT / "data.XLSX"))

# Where Client drops their Excel file if name not data.XLSX it will be converted to that path
DROP_DIR = Path(os.getenv("DATA_DROP_DIR") or PROJECT_ROOT)

# prefer names that match this regex
PREFER_NAME_REGEX = os.getenv("DATA_PREFER_REGEX") or ""

# How long to wait for file to upload
try:
    DATA_STABLE_SEC = int(os.getenv("DATA_STABLE_SEC") or "2")
except ValueError:
    DATA_STABLE_SEC = 2

# ==== Multi-Excel support: scenario / folder modes ====

# Root folder that holds scenario subfolders (default: <project>/scenarios)
SCENARIOS_ROOT = Path(os.getenv("SCENARIOS_ROOT") or (PROJECT_ROOT / "scenarios"))

# Optional “official” scenarios (only used if you call --scenario)
SCENARIO_FILES = {
    "P2T": [
        "OneTimePriorityJobs_P2T_Patch.xlsx",
        "OneTimeJobs_P2T_Patch.xlsx",
        "OneTimeJobs_P2T.xlsx",
        "ScheduledJobs_P2T.xlsx",
    ],
    "QuarterlyPatch": [
        "OneTimePriorityJobs_P2T_Patch.xlsx",
        "OneTimeJobs_P2T_Patch.xlsx",
        "OneTimeJobs_P2T.xlsx",
    ],
}

# --- Password-file override for scheduler (simple + backward-compatible) ---
PASSWORD_FILE = PROJECT_ROOT / "password" / "password.txt"   # one line: latest scheduler password

def _read_password_file() -> str | None:
    """Return password from ./password/password.txt if present & non-empty; else None."""
    try:
        if PASSWORD_FILE.exists():
            val = PASSWORD_FILE.read_text(encoding="utf-8").strip()
            return val or None
    except Exception:
        # never raise for password file; silently ignore and fall back to .env
        pass
    return None


def _parse_args():
    ap = argparse.ArgumentParser(add_help=False)
    ap.add_argument("--env", default="", help="Environment e.g TEST, DEV1, DEV2, ETC")
    ap.add_argument("--scenario", choices=SCENARIO_FILES.keys(),
                    help="Run a predefined file list (strict names) like P2T or QuarterlyPatch")
    ap.add_argument("--folder", help="Process all Excel files under scenarios/<folder> (any names)")
    ap.add_argument("--strict-order", action="store_true",
                    help="Folder mode only: process files in alphabetical order (default: newest first)")
    ap.add_argument("--help", action="help")
    return ap.parse_args()

def _existing_file_in_drop(name: str) -> Path | None:
    # Look in DROP_DIR first, then project root, then scenarios root
    for base in [DROP_DIR, PROJECT_ROOT, SCENARIOS_ROOT]:
        p = Path(base) / name
        if p.exists() and p.is_file() and not p.name.startswith("~$"):
            return p
    return None

def _collect_excel_paths(args) -> list[Path]:
    """Return a list of Excel file paths to process based on args.
       If no args provided, falls back to your current single-file behavior."""
    # 1) scenario mode — exact filenames (controlled runs)
    if args.scenario:
        out = []
        for fname in SCENARIO_FILES[args.scenario]:
            p = _existing_file_in_drop(fname)
            if p:
                out.append(p)
            else:
                logger.warning(f"[scenario:{args.scenario}] Missing file: {fname}")
        return out

    # 2) folder mode — any *.xls* under scenarios/<folder> (flexible runs)
    if args.folder:
        folder = SCENARIOS_ROOT / args.folder
        if not folder.exists():
            logger.error(f"Scenario folder not found: {folder}")
            return []
        cands = [p for p in folder.glob("*.xls*") if p.is_file() and not p.name.startswith("~$")]
        if not cands:
            logger.warning(f"No Excel files under {folder}")
            return []
        if args.strict_order:
            cands.sort(key=lambda p: p.name.lower())                  # predictable alpha
        else:
            cands.sort(key=lambda p: p.stat().st_mtime, reverse=True) # newest first
        return cands

    # 3) fallback — your existing single-drop behavior
    eff = refresh_data_xlsx_from_drop()
    return [eff] if eff else []


DRY_RUN = False  # True = print payloads, no network calls. False = real calls.

# read Oracle tenant & creds from .env. rstrip("/") avoids double slashes when we build URLs
BASEURL = get_env("FUSION_BASEURL", ENV_PREFIX).rstrip("/")
# USERNAME/PASSWORD are now chosen per Excel row; we only need BASEURL to be present.
if not BASEURL:
    raise SystemExit("Missing FUSION_BASEURL in .env")

# -------- Excel helpers --------
def _is_file_stable(path: Path, seconds: int) -> bool:
    try:
        s1 = path.stat().st_size
        time.sleep(seconds)
        s2 = path.stat().st_size
        return s1 == s2
    except FileNotFoundError:
        return False

def _pick_latest_excel(drop_dir: Path, prefer_regex: str = "") -> Path | None:
    cands = [
        p for p in drop_dir.glob("*")
        if p.is_file()
        and p.suffix.lower() in (".xlsx", ".xlsm")
        and not p.name.startswith("~$")
    ]
    if not cands:
        return None

    if prefer_regex:
        try:
            pat = re.compile(prefer_regex, re.IGNORECASE)
            preferred = [p for p in cands if pat.search(p.name)]
            if preferred:
                cands = preferred
        except re.error as e:
            logger.warning(f"Ignoring bad DATA_PREFER_REGEX: {e}")

    cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return cands[0]

def refresh_data_xlsx_from_drop() -> Path | None:
    try:
        latest = _pick_latest_excel(DROP_DIR, PREFER_NAME_REGEX)
        if not latest:
            if DATA_XLSX_PATH.exists():
                logger.info(f"No new Excel found; using existing {DATA_XLSX_PATH.name}.")
                return DATA_XLSX_PATH
            logger.warning(f"No Excel found in {DROP_DIR} and {DATA_XLSX_PATH} doesn’t exist.")
            return None

        if latest.resolve() == DATA_XLSX_PATH.resolve():
            if _is_file_stable(DATA_XLSX_PATH, DATA_STABLE_SEC):
                logger.info(f"Using existing {DATA_XLSX_PATH.name} (already latest).")
                return DATA_XLSX_PATH
            logger.warning(f"{DATA_XLSX_PATH.name} not yet stable; try again shortly.")
            return None

        if not _is_file_stable(latest, DATA_STABLE_SEC):
            logger.warning(f"Latest Excel not stable yet: {latest.name} (size still changing).")
            return None

        shutil.copy2(latest, DATA_XLSX_PATH)
        logger.info(f"Updated {DATA_XLSX_PATH.name} from drop file: {latest.name}")
        return DATA_XLSX_PATH
    except Exception as e:
        logger.exception(f"refresh_data_xlsx_from_drop failed: {e}")
        return None

def read_jobs_from_excel(path: Union[str, Path] = DATA_XLSX_PATH) -> List[Dict[str, str]]:
    path = Path(path)
    if not path.exists():
        logger.error(f"Excel not found at: {path}")
        return []

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_idx = {i: h for i, h in enumerate(headers) if h}

    job_rows: List[Dict[str, str]] = []
    for r in rows[1:]:
        row_dict: Dict[str, str] = {}
        is_empty = True
        for i, cell in enumerate(r):
            if i in header_idx:
                val = "" if cell is None else str(cell).strip()
                row_dict[header_idx[i]] = val
                if val:
                    is_empty = False
        if not is_empty:
            job_rows.append(row_dict)

    return job_rows

# ---------------- Param extraction (robust) ----------------
RESERVED_COLS = {
    "Display Name","jobDefinitionId","jobPackageName","jobDefinitionName","application",
    "startDate","endDate","icalString","schedule_json","className", "Username"
}
ARG_PREFIX = "argument"

_NULL_SENTINELS = {"null", "none", "n/a", "na"}  # treated as empty

def _nullish(v: str | None) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return (s == "") or (s.lower() in _NULL_SENTINELS)

# normalize headers like "Argument 2", "ARGUMENT2", "argument  2" → "argument2"
_ARG_HEADER_RX = re.compile(r"^argument\s*(\d{1,2})$", re.IGNORECASE)
def _normalize_arg_header(h: str) -> str | None:
    m = _ARG_HEADER_RX.match(h.strip())
    return f"argument{m.group(1)}" if m else None

def extract_params(row: Dict[str, str]) -> Dict[str, str]:
    """
    Robust parameter collection:
      - Named prompts are kept EXACT (e.g., 'Business Unit', 'submit.argument1') unless they are reserved.
      - Values that are '', 'null', 'None', 'N/A' are treated as empty and skipped.
      - argument1..30 accepted case/space-insensitive (e.g., 'Argument 2' → 'argument2').
    """
    params: Dict[str, str] = {}

    # 1) Named prompts
    for k, v in row.items():
        if _nullish(v):
            continue
        if not k:
            continue
        if k in RESERVED_COLS:
            continue
        if _normalize_arg_header(k):
            continue
        params[k.strip()] = str(v).strip()

    # 2) Generic arguments (argument1..30)
    for n in range(1, 31):
        key = f"argument{n}"
        raw = row.get(key, "")
        if _nullish(raw):
            for hdr, val in row.items():
                if _normalize_arg_header(hdr) == key:
                    raw = val
                    break
        if _nullish(raw):
            continue
        params[key] = str(raw).strip()

    return params

def extract_schedule(row: Dict[str, str]) -> Dict:
    sch_json = (row.get("schedule_json") or "").strip()
    if sch_json:
        obj = json.loads(sch_json)
        if not isinstance(obj, dict):
            raise ValueError("schedule_json must be a JSON object")
        return obj

    ical = (row.get("icalString") or "").strip()
    start_date = (row.get("startDate") or "").strip()
    end_date = (row.get("endDate") or "").strip()
    if ical:
        rec: Dict[str, str] = {"icalString": ical}
        if start_date:
            rec["startDate"] = start_date
        if end_date:
            rec["endDate"] = end_date
        return {"recurrences": [rec]}
    return {}

def is_on_demand(schedule: Dict) -> bool:
    return not schedule

def load_accounts_from_env() -> dict:
    """
    Build {login_email_lower: {"baseurl":..., "username":..., "password":...,}} from .env only.
    Optional password-file override (scheduler only):
      FUSION_SCHEDULER_LOGIN=<login to override>
      -> if set and ./password/password.txt exists/non-empty, that password replaces the .env password.
    """
    base = get_env("FUSION_BASEURL", ENV_PREFIX).rstrip("/")
    aliases = [a.strip() for a in (get_env("FUSION_USERS", ENV_PREFIX) or "").split(",") if a.strip()]
    if not base or not aliases:
        raise SystemExit("Missing FUSION_BASEURL or FUSION_USERS in .env")

    scheduler_login_env = get_env("FUSION_SCHEDULER_LOGIN", ENV_PREFIX).strip().lower()
    out = {}

    for alias in aliases:
        login = get_env(f"FUSION_{alias}_LOGIN", ENV_PREFIX).strip()
        pwd   = get_env(f"FUSION_{alias}_PASSWORD", ENV_PREFIX).strip()
        if not login:
            raise SystemExit(f"Missing login for alias {alias}: set FUSION_{alias}_LOGIN")

        # Apply override if this alias matches scheduler login
        if scheduler_login_env and login.lower() == scheduler_login_env:
            file_pwd = _read_password_file()
            if file_pwd:
                pwd = file_pwd
                logger.info(f"[auth] Using password file for {login} (source=file)")
            else:
                logger.info(f"[auth] Using .env password for {login} (source=env)")

        if not pwd:
            raise SystemExit(
                f"Missing password for alias {alias}: set FUSION_{alias}_PASSWORD "
                f"or provide password/password.txt if this alias is the scheduler."
            )

        out[login.lower()] = {"baseurl": base, "username": login, "password": pwd}
    return out

def pick_login_from_row(row: dict) -> str:
    #requires username column with actual username
    for k in row.keys():
        if k and k.strip().lower() == "username":
            v = (row.get(k) or "").strip()
            if not v:
                raise ValueError("Excel row missing Username login")
            return v
    raise ValueError("Excel is missing a 'Username' column header")

def build_client_cache_env_only():
    accounts_by_login = load_accounts_from_env()
    cache = {}
    def get(login_email: str):
        key = login_email.lower().strip()
        if key not in accounts_by_login:
            known = ", ".join(accounts_by_login.keys())
            raise ValueError(f"Unknown Username '{login_email}'. Known: {known}")
        if key not in cache:
            cfg = accounts_by_login[key]
            cache[key] = ESSClient(cfg["baseurl"], cfg["username"], cfg["password"])
        return cache[key]
    return get

# ------------------------- MAIN RUN FUNCTION -------------------------
def run() -> int:
    # NEW: parse optional args
    args = _parse_args()

    # NEW: collect files to process (scenario / folder / fallback single)
    excel_paths = _collect_excel_paths(args)
    if not excel_paths:
        logger.info("No Excel files found to process.")
        return 0

    get_client = build_client_cache_env_only()  # unchanged: still uses .env creds (+password file override)
    total_rows = 0

    for xlsx in excel_paths:
        try:
            # Stage each Excel as data.XLSX so your existing reader path stays the same
            if xlsx.resolve() != DATA_XLSX_PATH.resolve():
                if not _is_file_stable(xlsx, DATA_STABLE_SEC):
                    logger.warning(f"File not yet stable (skip this run): {xlsx.name}")
                    continue
                shutil.copy2(xlsx, DATA_XLSX_PATH)
                logger.info(f"Using {xlsx.name} → staged as {DATA_XLSX_PATH.name}")
            else:
                logger.info(f"Using existing {DATA_XLSX_PATH.name}.")

            rows = read_jobs_from_excel(DATA_XLSX_PATH)
            if not rows:
                logger.warning(f"No rows found in {xlsx.name}")
                continue

            for i, row in enumerate(rows, start=1):
                login = pick_login_from_row(row)   
                client = get_client(login)
                logger.info(f"[{login}] {xlsx.name} row {i} selected account")

                job_definition_id   = (row.get("jobDefinitionId") or "").strip()
                job_package_name    = (row.get("jobPackageName") or "").strip().rstrip("/")
                job_definition_name = (row.get("jobDefinitionName") or "").strip()
                application         = (row.get("application") or "").strip()
                description         = (row.get("Display Name") or "").strip()
                class_name          = (row.get("className") or "").strip()
                display_name        = description

                # Resolve job by display name if necessary
                if (not job_definition_id and not (job_package_name and job_definition_name)) or not application:
                    try:
                        resolved = client.resolve_job_by_name(display_name)
                    except Exception:
                        resolved = None
                    if resolved:
                        application = application or resolved.get("application") or application
                        if not (job_package_name and job_definition_name):
                            job_package_name    = job_package_name or resolved.get("jobPackageName") or ""
                            job_definition_name = job_definition_name or resolved.get("jobDefinitionName") or ""
                            job_definition_id   = job_definition_id or resolved.get("jobDefinitionId") or job_definition_id
                        logger.info(
                            f"Resolved '{display_name}' → app={application or '-'} "
                            f"pkg={job_package_name or '-'} name={job_definition_name or '-'}"
                        )
                    else:
                        logger.error(
                            f"Row {i}: cannot resolve jobDefinition fields from Display Name '{display_name}'. Skipping."
                        )
                        continue

                # --- Prefer jobDefinitionId path over package+name to avoid SYS_AdHocRequest (403) ---
                if job_package_name and job_definition_name and not job_definition_id:
                    job_definition_id = f"JobDefinition://{job_package_name.lstrip('/')}/{job_definition_name}"
                if job_definition_id:
                    job_package_name = ""
                    job_definition_name = ""

                # params (kept exactly as your current logic)
                params = extract_params(row)

                # Mirror generic args onto submit.argumentN (and vice versa)
                for n in range(1, 30 + 1):
                    a  = f"argument{n}"
                    sa = f"submit.argument{n}"
                    if params.get(a) and not params.get(sa):
                        params[sa] = params[a]
                    if params.get(sa) and not params.get(a):
                        params[a] = params[sa]

                # OSCS alias fan-out (unchanged)
                def _is_oscs_job() -> bool:
                    jdid = (job_definition_id or "").lower()
                    pkg  = (job_package_name or "").lower()
                    name = (job_definition_name or "").lower()
                    disp = (display_name or "").lower()
                    return (
                        "oscs" in disp
                        or "fndoscs" in name
                        or "/fnd/applcore/" in jdid
                        or "/fnd/applcore/" in pkg
                    )

                if _is_oscs_job():
                    OSCS_PARAM_ALIASES = [
                        "Index Name to Reingest",
                        "Index Name for recreate",
                        "indexName",
                        "argument1",
                        "submit.argument1",
                    ]
                    ix_val = next((params.get(k) for k in OSCS_PARAM_ALIASES if params.get(k)), None)
                    if ix_val:
                        for key in ("Index Name for recreate", "indexName", "argument1", "submit.argument1"):
                            params.setdefault(key, ix_val)

                if not params:
                    logger.warning(
                        f"Row {i}: no parameters extracted for '{display_name}'. "
                        f"If the job requires prompts, ensure Excel headers match exactly "
                        f"(e.g., 'Business Unit', 'Ledger') or use 'argument1..30'."
                    )

                # Schedule extraction / validation
                try:
                    schedule = extract_schedule(row)
                except Exception as e:
                    logger.exception(f"Row {i}: invalid schedule: {e}")
                    continue

                logger.info(f"Row {i}: parameters (dict) → {json.dumps(params, ensure_ascii=False)}")

                # duplicate / existing handling (unchanged)
                existing_id = None
                row_start_date = (row.get("startDate") or "").strip() #new**********
                row_argument1  = params.get("argument1", "").strip() #new************
                if job_definition_id:
                    try:
                        if is_on_demand(schedule): #new ***************
                            existing_id = client.find_active_request_for_definition(job_definition_id)
                        else: # SCHEDULED: smart check — all three must match to be duplicate new***********
                            existing_id = client.find_active_scheduled_request(
                                job_definition_id,
                                start_date=row_start_date,
                                argument1=row_argument1,
                            ) #this section new**********
                    except Exception:
                        existing_id = None

                if existing_id:
                    logger.info(
                        f"Existing active request for {job_definition_id}: request_id={existing_id} — not submitting a duplicate."
                    )
                    if is_on_demand(schedule):
                        if DRY_RUN:
                            logger.info("[DRY RUN] Would poll the existing request to completion (skipped).")
                        else:
                            final = client.poll_until_complete(
                                existing_id,
                                poll_seconds=int(get_env("DUPLICATE_POLL_INTERVAL_SECONDS", ENV_PREFIX) or "10"),
                                timeout_seconds=int(get_env("DUPLICATE_POLL_TIMEOUT_SECONDS", ENV_PREFIX) or "300"),
                                logger=logger,
                            )
                            final_status = final.get("Status") or final.get("status") or final
                            logger.info(
                                f"[ON-DEMAND EXISTING] request_id={existing_id} def={job_definition_id} final_status={final_status}"
                            )
                            # >>> AUDIT <<< existing on-demand, final status
                            try:
                                job_name_for_audit = job_definition_name or display_name or "UnknownJob"
                                log_job_run_xlsx(
                                    job_name=job_name_for_audit,
                                    request_id=str(existing_id),
                                    status=str(final_status),
                                    fusion_alias=str(login),
                                    notes="existing_on_demand"
                                )
                            except Exception:
                                pass
                    else:
                        logger.info(f"[SCHEDULED] Another schedule/run already active for def={job_definition_id} startDate={row_start_date} argument1={row_argument1 or '-'}; skipping submit.") #this is new one**********
                        #logger.info(f"[SCHEDULED] Another schedule/run already active for def={job_definition_id}; skipping submit.")   this is old one
                    continue

                # submit (unchanged)
                try:
                    req_id = client.submit(
                        job_definition_id,
                        application,
                        params,
                        schedule if not is_on_demand(schedule) else {},
                        description=description,
                        dry_run=DRY_RUN,
                        logger=logger,
                        job_package_name=job_package_name or None,
                        job_definition_name=job_definition_name or None,
                        class_name=class_name or None,
                    )

                    if DRY_RUN:
                        if is_on_demand(schedule):
                            logger.info("[DRY RUN] Would poll until completion for on-demand job (skipped).")
                        else:
                            logger.info("[DRY RUN] Would log scheduled parent requestId (skipped).")
                        continue

                    if is_on_demand(schedule):
                        final = client.poll_until_complete(req_id, logger=logger)
                        final_status = final.get("Status") or final.get("status") or final
                        logger.info(
                            f"[ON-DEMAND] request_id={req_id} pkg={job_package_name or '-'} "
                            f"name={job_definition_name or '-'} final_status={final_status}"
                        )
                        # >>> AUDIT <<< on-demand, final status
                        try:
                            job_name_for_audit = job_definition_name or display_name or "UnknownJob"
                            log_job_run_xlsx(
                                job_name=job_name_for_audit,
                                request_id=str(req_id),
                                status=str(final_status),
                                fusion_alias=str(login),
                                notes="on_demand"
                            )
                        except Exception:
                            pass
                    else:
                        logger.info(
                            f"[SCHEDULED] parent_request_id={req_id} pkg={job_package_name or '-'} "
                            f"name={job_definition_name or '-'} schedule_created"
                        )
                        # >>> AUDIT <<< scheduled parent request id
                        try:
                            job_name_for_audit = job_definition_name or display_name or "UnknownJob"
                            log_job_run_xlsx(
                                job_name=job_name_for_audit,
                                request_id=str(req_id),
                                status="SCHEDULED",
                                fusion_alias=str(login),
                                notes="scheduled_parent"
                            )
                        except Exception:
                            pass

                except ESSDuplicatePendingError:
                    logger.warning(f"ESS-01050: another request is already pending for {job_definition_id}.")
                    continue
                except Exception as e:
                    logger.exception(f"Row {i}: submit failed: {e}")

                total_rows += 1

        except Exception as e:
            logger.exception(f"Failed processing file: {xlsx.name} → {e}")

    logger.info(f"Processed {total_rows} rows across {len(excel_paths)} file(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
