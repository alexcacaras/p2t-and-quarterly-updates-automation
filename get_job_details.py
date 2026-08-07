import sys
from pathlib import Path
from typing import Optional, Union

import requests
from dotenv import dotenv_values
from openpyxl import Workbook, load_workbook

from fusion_api import JobRequest

# Load environment values from ".env"
config = dotenv_values()

# Assign environment variables
baseurl = config["FUSION_BASEURL"]
username = config["FUSION_USERNAME"] or ""
password = config["FUSION_PASSWORD"] or ""


def get_job_request_info(request_id: str) -> Optional[JobRequest]:
    """_summary_

    Args:
        request_id (str): _description_

    Returns:
        Optional[JobRequest]: _description_
    """
    response = requests.get(
        f"{baseurl}/ess/rest/scheduler/v1/requests/{request_id}?fields=@full",
        auth=(username, password),
    )
    if not response.ok:
        print(f"{response.status_code} {response.reason}")
        return

    return response.json()


def format_row(job_request_info: JobRequest) -> list[str]:
    return ["a"]


def write_to_excel(filename: Union[str, Path], job_request_info: JobRequest) -> None:
    """_summary_

    Args:
        filename (Union[str, Path]): _description_
        job_request_info (JobRequest): _description_
    """
    column_headers = [
        "Display Name",
        "jobDefinitionId",
        "application",
        "startDate",
        "endDate",
        "icalString",
        "argument1",
        "argument2",
        "argument3",
        "argument4",
        "argument5",
        "argument6",
        "argument7",
        "argument8",
        "argument9",
        "argument10",
        "argument11",
        "argument12",
        "argument13",
        "argument14",
        "argument15",
        "argument16",
        "argument17",
        "argument18",
        "argument19",
        "argument20",
        "argument21",
        "argument22",
        "argument23",
        "argument24",
        "argument25",
        "argument26",
        "argument27",
        "argument28",
        "argument29",
        "argument30",
    ]
    file_path = Path(filename)
    if file_path.exists():
        # File exists, append data
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()

    worksheet = workbook.active
    if worksheet is None:
        raise ValueError(
            f"No active worksheet. Ensure that the workbook at {filename} has an active worksheet."
        )
    if not file_path.exists():
        worksheet.append(column_headers)

    workbook.save(file_path)
    print(f"Data successfully writtent to file {filename}.")


def main(request_id: str, filename: str = "ESSjobs.xlsx"):
    job_request_info = get_job_request_info(request_id)
    if job_request_info:
        write_to_excel(filename, job_request_info)


if __name__ == "__main__":
    if len(sys.argv) == 2:
        main(sys.argv[1])
    elif len(sys.argv) == 3:
        main(sys.argv[1], sys.argv[2])
    else:
        print("Please provide the request id as an arguement.")
