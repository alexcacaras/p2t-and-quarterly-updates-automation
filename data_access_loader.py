from typing import Any, Dict, List            # type hints for return values
import requests                               # HTTP requests to Oracle Fusion REST API
import os                                     # read environment variables
from openpyxl import load_workbook            # read Data Access Excel file
from requests.auth import HTTPBasicAuth       # basic auth for API session
from crypto_env import load_env_from_encrypted, get_env, _get_env_prefix  # loads credentials from encrypted .env file into memory
from pathlib import Path                      # handle file paths cross-platform
load_env_from_encrypted()                     # decrypt and inject .env vars into os.environ at module load

ENV_PREFIX = _get_env_prefix()
filepath = "DATA_ACCESS/Data Access.xlsx"


class DataAccessClient:
    """
    Flow:
    1) GET existing dataSecurities to build a map of display name → internal role code
    2) Read the Data Access Excel from DATA_ACCESS folder
    3) Translate display names to internal role codes using:
         - Col H (manual code) if provided → use directly
         - Role map lookup from API → use if found
         - Fallback → use Col F as-is
    4) POST each record to /fscmRestApi/resources/11.13.18.05/dataSecurities

    Excel columns:
        Col A — SecurityContext
        Col B — SecurityContextValue
        Col E — UserName
        Col F — Role display name (e.g. "Accounts Payable Manager")
        Col G — Active (Yes/No)
        Col H — Role code override (optional, e.g. "ORA_AP_MANAGER_JOB")
                 If empty, role map lookup is used automatically. If COL H added will be faster as the mapping takes about 2min.

    Required privilege on auth account:
        Manage Data Access
    Note:
        I think if accounts getting the data accessdo not have the role for their specific data access will get error message like this RESPONSE 400: The value of the attribute Role isn't valid.
    """

    def __init__(self):
        self.base_url = get_env("FUSION_BASEURL", ENV_PREFIX).rstrip("/")
        self.session = requests.Session()
        self.session.auth = HTTPBasicAuth(
            get_env("FUSION_it_LOGIN", ENV_PREFIX),
            get_env("FUSION_it_PASSWORD", ENV_PREFIX)
        )

    # --------------------------------------------
    # Step 0a - Build Role Map from API
    # --------------------------------------------

    def _build_role_map(self) -> Dict[str, str]:
        """
        GET all existing dataSecurities records and build a map of:
        display name → RoleCommonName (internal code)
        e.g. "Accounts Payable Manager" → "ORA_AP_MANAGER_JOB"

        Paginates through all records using limit/offset.
        Falls back to display name if role not found in map.
        """
        url = f"{self.base_url}/fscmRestApi/resources/11.13.18.05/dataSecurities"
        role_map = {}
        offset = 0

        print("  → Building role map from Oracle...")

        while True:
            resp = self.session.get(url, params={"limit": 500, "offset": offset}, timeout=60)
            resp.raise_for_status()
            data = resp.json()

            for item in data.get("items", []):
                display = item.get("RoleNameCr", "").strip()
                code    = item.get("RoleCommonName", "").strip()
                if display and code:
                    role_map[display] = code

            if not data.get("hasMore"):
                break
            offset += 500

        print(f"  → Role map built: {len(role_map)} unique roles found")
        return role_map

    # --------------------------------------------
    # Step 0b - Read Excel Sheet
    # --------------------------------------------

    def read_from_excel(self, filepath: str) -> List[Dict[str, Any]]:
        """
        Read data access records from Excel.
        Col H (manual role code) is optional — if empty, role map is used in run().
        """
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        records = []
        for row in rows[1:]:  # skip header
            if not any(row):  # skip empty rows
                continue

            # Col H — manual role code override (optional)
            manual_code = ""
            if len(row) > 7 and row[7] is not None:
                manual_code = str(row[7]).strip()

            records.append({
                "SecurityContext":      str(row[0]).strip(),  # Col A
                "SecurityContextValue": str(row[1]).strip(),  # Col B
                "UserName":             str(row[4]).strip(),  # Col E
                "RoleCommonName":       str(row[5]).strip(),  # Col F — display name, translated in run()
                "ActiveFlag":           str(row[6]).strip().lower() == "yes",  # Col G
                "_manual_code":         manual_code           # Col H — internal, removed before POST
            })
        return records

    # ---------------------------------------------
    # Step 1 - POST Data Access Record
    # ---------------------------------------------

    def post_data_access(self, record: Dict[str, Any]) -> Dict:
        """
        POST a single data access record to dataSecurities endpoint.
        Skips if record already exists (duplicate).
        """
        url = f"{self.base_url}/fscmRestApi/resources/11.13.18.05/dataSecurities"

        resp = self.session.post(url, json=record, timeout=120)

        if resp.status_code != 201:
            print(f"  RESPONSE {resp.status_code}: {resp.text}")

        if resp.status_code == 400 and "already exists" in resp.text:
            print(f"  → Skipped (already exists): {record['UserName']} / {record['RoleCommonName']}")
            return {"status": "skipped", "record": record}

        resp.raise_for_status()
        print(f"  → Posted: {record['UserName']} / {record['RoleCommonName']}")
        return {"status": "posted", "record": record}

    # ----------------------------------------
    # Step 2 - Run
    # ----------------------------------------

    def run(self, dry_run: bool = False) -> List[Dict]:
        """
        Full data access load flow:
        1) Build role map from API
        2) Read records from Excel
        3) Translate display names to internal role codes
           Priority: Col H manual code → role map → fallback to Col F as-is
        4) POST each record (or dry run)
        """
        results = []

        print(f"\n{'='*50}")
        print("DATA ACCESS LOAD")
        print(f"{'='*50}")

        # Step 0a — read records from Excel first
        records = self.read_from_excel(filepath)
        print(f"  → Found {len(records)} records to process")

        # Step 0b — only build role map if any row is missing Col H code
        needs_map = any(not r["_manual_code"] for r in records)
        if needs_map:
            role_map = self._build_role_map()
        else:
            role_map = {}
            print("  → All rows have manual codes, skipping role map build")

        # Step 1 — translate display names → internal codes
        for record in records:
            manual_code  = record.pop("_manual_code")  # remove internal field before POST
            display_name = record["RoleCommonName"]

            if manual_code:
                # Col H has a code — use it directly
                record["RoleCommonName"] = manual_code
                print(f"  → Using manual code for '{display_name}': {manual_code}")
            else:
                # Col H empty — try role map lookup
                internal_code = role_map.get(display_name)
                if internal_code:
                    record["RoleCommonName"] = internal_code
                else:
                    print(f"  Role not found in map, using as-is: '{display_name}'")

        # Step 2 — POST each record
        for record in records:
            try:
                if dry_run:
                    print(f"  [DRY RUN] Would POST: {record['UserName']} / {record['RoleCommonName']}")
                    results.append({"status": "dry_run", "record": record})
                else:
                    result = self.post_data_access(record)
                    results.append(result)
            except Exception as e:
                print(f"  ERROR: {record['UserName']} / {record['RoleCommonName']} - {e}")
                results.append({"status": "failed", "record": record, "error": str(e)})

        # Step 3 — summary
        posted  = sum(1 for r in results if r["status"] == "posted")
        skipped = sum(1 for r in results if r["status"] == "skipped")
        failed  = sum(1 for r in results if r["status"] == "failed")
        dry     = sum(1 for r in results if r["status"] == "dry_run")
        print(f"\n  Done | Posted: {posted} | Skipped: {skipped} | Failed: {failed} | Dry Run: {dry}")

        return results


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--env", default="")
    ap.parse_args()
    client = DataAccessClient()
    results = client.run(dry_run=False)
    print(f"\nResults: {results}")